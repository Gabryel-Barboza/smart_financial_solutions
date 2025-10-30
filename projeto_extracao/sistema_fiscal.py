# sistema_fiscal.py
import os
import logging
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

import pandas as pd
import xml.etree.ElementTree as ET
import zipfile
import tempfile
import re
import requests
import json
from datetime import datetime
from sqlalchemy import create_engine, Column, String, Integer, Float, DateTime, Text, ForeignKey, JSON, Index
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from sqlalchemy import text
from typing import Dict, List, Any, Optional
import warnings
from openpyxl import load_workbook

try:
    from config import MYSQL_CONFIG, get_mysql_uri
except ImportError as e:
    logger.error(f"Erro ao importar configurações: {e}")
    MYSQL_CONFIG = {
        'host': 'localhost',
        'port': 3306,
        'user': 'root',
        'password': '',
        'database': 'fiscal_db'
    }
    
    def get_mysql_uri():
        return f"mysql+pymysql://{MYSQL_CONFIG['user']}:{MYSQL_CONFIG['password']}@{MYSQL_CONFIG['host']}:{MYSQL_CONFIG['port']}/{MYSQL_CONFIG['database']}"

warnings.filterwarnings('ignore')
Base = declarative_base()

# =========================================================
#   MODELOS DE BANCO DE DADOS
# =========================================================

class Emitente(Base):
    __tablename__ = "emitentes"
    
    id_emitente = Column(Integer, primary_key=True, autoincrement=True)
    cnpj = Column(String(14), unique=True, nullable=False, index=True)
    razao_social = Column(String(255))
    nome_fantasia = Column(String(255))
    ie = Column(String(20))
    logradouro = Column(String(255))
    numero = Column(String(10))
    complemento = Column(String(100))
    bairro = Column(String(100))
    municipio = Column(String(100))
    uf = Column(String(2))
    cep = Column(String(8))
    pais = Column(String(50), default="Brasil")
    telefone = Column(String(20))
    created_at = Column(DateTime, default=datetime.now)

class Destinatario(Base):
    __tablename__ = "destinatarios"
    
    id_destinatario = Column(Integer, primary_key=True, autoincrement=True)
    cnpj = Column(String(14), unique=True, nullable=False, index=True)
    razao_social = Column(String(255))
    ie = Column(String(20))
    logradouro = Column(String(255))
    numero = Column(String(10))
    complemento = Column(String(100))
    bairro = Column(String(100))
    municipio = Column(String(100))
    uf = Column(String(2))
    cep = Column(String(8))
    pais = Column(String(50), default="Brasil")
    telefone = Column(String(20))
    created_at = Column(DateTime, default=datetime.now)

class NotaFiscal(Base):
    __tablename__ = "notas_fiscais"
    
    id_nf = Column(Integer, primary_key=True, autoincrement=True)
    numero = Column(String(15), nullable=False, index=True)
    serie = Column(String(3))
    modelo = Column(String(10), nullable=True)  # Modelo da NF (55, 65, etc.)
    data_emissao = Column(DateTime)
    data_saida = Column(DateTime)
    natureza_operacao = Column(String(100))
    tipo_operacao = Column(String(50))
    finalidade_nf = Column(String(50))
    valor_total = Column(Float)
    valor_produtos = Column(Float)
    valor_icms = Column(Float)
    valor_pis = Column(Float)
    valor_cofins = Column(Float)
    
  
    
    # Relacionamentos
    cnpj_emitente = Column(String(14), ForeignKey('emitentes.cnpj'), nullable=False, index=True)
    razao_emitente = Column(String(255))
    cnpj_destinatario = Column(String(14), ForeignKey('destinatarios.cnpj'), nullable=True, index=True)
    razao_destinatario = Column(String(255))
    
    codigo_municipio = Column(String(7))
    uf_emitente = Column(String(2))
    uf_destinatario = Column(String(2))
    
    # Relacionamentos ORM
    emitente_rel = relationship("Emitente", foreign_keys=[cnpj_emitente])
    destinatario_rel = relationship("Destinatario", foreign_keys=[cnpj_destinatario])
    
    chave_acesso = Column(String(44), unique=True, index=True)
    layout_detectado = Column(String(100))
    tipo_arquivo = Column(String(20))
    created_at = Column(DateTime, default=datetime.now)
    
    # Índice composto para busca eficiente
    __table_args__ = (
        Index('idx_notas_cnpj_numero', 'cnpj_emitente', 'numero'),
        Index('idx_notas_chave', 'chave_acesso'),
    )

class ItemNF(Base):
    __tablename__ = "itens_nf"
    
    id_item = Column(Integer, primary_key=True, autoincrement=True)
    id_nf = Column(Integer, ForeignKey('notas_fiscais.id_nf'), nullable=False, index=True)
    n_item = Column(Integer, nullable=False)
    codigo_produto = Column(String(60))
    descricao_produto = Column(Text)
    cfop = Column(String(10))
    quantidade = Column(Float)
    unidade = Column(String(10))
    valor_unitario = Column(Float)
    valor_total = Column(Float)
    aliquota_icms = Column(Float)
    valor_icms = Column(Float)
    aliquota_pis = Column(Float)
    valor_pis = Column(Float)
    aliquota_cofins = Column(Float)
    valor_cofins = Column(Float)
    created_at = Column(DateTime, default=datetime.now)
    
    # Relacionamento
    nota_fiscal = relationship("NotaFiscal", backref="itens")
    
    __table_args__ = (
        Index('idx_itens_nf', 'id_nf', 'n_item'),
    )

class TotalNF(Base):
    __tablename__ = "totais_nf"
    
    id_total = Column(Integer, primary_key=True, autoincrement=True)
    id_nf = Column(Integer, ForeignKey('notas_fiscais.id_nf'), nullable=False, unique=True)
    base_icms = Column(Float)
    valor_icms = Column(Float)
    base_icms_st = Column(Float)
    valor_icms_st = Column(Float)
    valor_produtos = Column(Float)
    valor_frete = Column(Float)
    valor_seguro = Column(Float)
    valor_desconto = Column(Float)
    valor_outros = Column(Float)
    valor_pis = Column(Float)
    valor_cofins = Column(Float)
    valor_nf = Column(Float)
    created_at = Column(DateTime, default=datetime.now)
    
    # Relacionamento
    nota_fiscal = relationship("NotaFiscal", backref="totais")

class TransporteNF(Base):
    __tablename__ = "transporte_nf"
    
    id_transporte = Column(Integer, primary_key=True, autoincrement=True)
    id_nf = Column(Integer, ForeignKey('notas_fiscais.id_nf'), nullable=False, unique=True)
    modalidade_frete = Column(String(50))
    cnpj_transportador = Column(String(14))
    razao_transportador = Column(String(255))
    placa_veiculo = Column(String(7))
    uf_veiculo = Column(String(2))
    rntrc = Column(String(20))
    valor_frete = Column(Float)
    created_at = Column(DateTime, default=datetime.now)
    
    # Relacionamento
    nota_fiscal = relationship("NotaFiscal", backref="transporte")

class EntregaRetiradaNF(Base):
    __tablename__ = "entrega_retirada_nf"
    
    id_local = Column(Integer, primary_key=True, autoincrement=True)
    id_nf = Column(Integer, ForeignKey('notas_fiscais.id_nf'), nullable=False)
    tipo_local = Column(String(20))
    cnpj = Column(String(14))
    logradouro = Column(String(255))
    numero = Column(String(10))
    complemento = Column(String(100))
    bairro = Column(String(100))
    municipio = Column(String(100))
    uf = Column(String(2))
    created_at = Column(DateTime, default=datetime.now)
    
    # Relacionamento
    nota_fiscal = relationship("NotaFiscal", backref="locais")

class LayoutTemplate(Base):
    __tablename__ = "layout_templates"
    
    id = Column(String(36), primary_key=True)
    nome_layout = Column(String(100))
    cnpj_empresa = Column(String(14))
    campos_template = Column(JSON)
    accuracy = Column(Float)
    created_at = Column(DateTime, default=datetime.now)

# =========================================================
#  AGENTE INTELIGENTE COM DEEPSEEK
# =========================================================

class AgenteInteligenteDeepSeek:
    def __init__(self):
        self.api_key = os.getenv('DEEPSEEK_API_KEY')
        self.base_url = "https://api.deepseek.com/v1/chat/completions"
        
    def analisar_dados_fiscais(self, dados: Dict[str, Any], contexto: str = "nota_fiscal") -> Dict[str, Any]:
        """
        Usa DeepSeek para análise inteligente dos dados fiscais
        """
        if not self.api_key:
            logger.warning("DeepSeek API key não configurada")
            return {"analisado": False, "erro": "API key não configurada"}
        
        try:
            prompt = self._criar_prompt_analise(dados, contexto)
            
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}"
            }
            
            payload = {
                "model": "deepseek-chat",
                "messages": [
                    {
                        "role": "system",
                        "content": "Você é um especialista em análise fiscal brasileira. Analise os dados de notas fiscais e forneça insights valiosos."
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                "temperature": 0.3,
                "max_tokens": 2000
            }
            
            response = requests.post(self.base_url, headers=headers, json=payload, timeout=60)
            response.raise_for_status()
            
            resultado = response.json()
            analise = resultado['choices'][0]['message']['content']
            
            return {
                "analisado": True,
                "analise": analise,
                "recomendacoes": self._extrair_recomendacoes(analise),
                "timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Erro na análise com DeepSeek: {e}")
            return {"analisado": False, "erro": str(e)}
    
    def validar_dados_fiscais(self, dados: Dict[str, Any]) -> Dict[str, Any]:
        """
        Valida dados fiscais usando DeepSeek
        """
        if not self.api_key:
            return {"validado": False, "erro": "API key não configurada"}
        
        try:
            prompt = f"""
            Valide os seguintes dados de nota fiscal contra as regras fiscais brasileiras:
            
            DADOS DA NOTA:
            {json.dumps(dados, indent=2, ensure_ascii=False)}
            
            Verifique:
            1. CNPJ do emitente e destinatário são válidos
            2. Valores numéricos são consistentes
            3. CFOP é apropriado para a operação
            4. Campos obrigatórios estão presentes
            5. Inconsistências nos valores totais
            
            Retorne um JSON com:
            - validado: boolean
            - problemas: lista de problemas encontrados
            - recomendacoes: lista de recomendações
            - severidade: "baixa", "media" ou "alta"
            """
            
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}"
            }
            
            payload = {
                "model": "deepseek-chat",
                "messages": [
                    {
                        "role": "system",
                        "content": "Você é um validador fiscal especializado em notas fiscais brasileiras. Retorne apenas JSON."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                "temperature": 0.1,
                "max_tokens": 1500
            }
            
            response = requests.post(self.base_url, headers=headers, json=payload, timeout=60)
            response.raise_for_status()
            
            resultado = response.json()
            validacao_texto = resultado['choices'][0]['message']['content']
            
            # Tentar parsear o JSON da resposta
            try:
                validacao = json.loads(validacao_texto)
                validacao["timestamp"] = datetime.now().isoformat()
                return validacao
            except:
                return {
                    "validado": False,
                    "problemas": ["Não foi possível parsear a resposta do validador"],
                    "recomendacoes": ["Verificar manualmente os dados"],
                    "severidade": "media",
                    "timestamp": datetime.now().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Erro na validação com DeepSeek: {e}")
            return {"validado": False, "erro": str(e)}
    
    def _criar_prompt_analise(self, dados: Dict[str, Any], contexto: str) -> str:
        """Cria prompt para análise baseado no contexto"""
        
        if contexto == "nota_fiscal":
            return f"""
            Analise esta nota fiscal e forneça insights:

            DADOS PRINCIPAIS:
            - Emitente: {dados.get('emitente', {}).get('razao_social', 'N/A')} (CNPJ: {dados.get('emitente', {}).get('cnpj', 'N/A')})
            - Destinatário: {dados.get('destinatario', {}).get('razao_social', 'N/A')} (CNPJ: {dados.get('destinatario', {}).get('cnpj', 'N/A')})
            - Valor Total: R$ {dados.get('totais', {}).get('valor_total_nota', 0):.2f}
            - Data Emissão: {dados.get('identificacao', {}).get('data_emissao', 'N/A')}
            - Natureza Operação: {dados.get('identificacao', {}).get('natureza_operacao', 'N/A')}

            ITENS: {len(dados.get('itens', []))} itens

            Forneça análise sobre:
            1. Padrões nos valores e quantidades
            2. Consistência dos dados
            3. Possíveis riscos fiscais
            4. Recomendações para melhoria
            5. Insights sobre o negócio baseado nos itens
            """
        
        elif contexto == "lote_processamento":
            return f"""
            Analise este lote de processamento fiscal:

            ESTATÍSTICAS:
            {json.dumps(dados, indent=2, ensure_ascii=False)}

            Forneça insights sobre:
            1. Eficiência do processamento
            2. Padrões nos erros
            3. Recomendações para otimização
            4. Previsão de volume futuro
            """
        
        else:
            return f"""
            Analise estes dados fiscais:

            {json.dumps(dados, indent=2, ensure_ascii=False)}

            Forneça insights relevantes para o contexto: {contexto}
            """
    
    def _extrair_recomendacoes(self, analise: str) -> List[str]:
        """Extrai recomendações específicas da análise"""
        recomendacoes = []
        
        # Padrões comuns de recomendações
        padroes = [
            r"Recomenda[çc][aã]o\s*[:\-]\s*(.+)",
            r"Sugest[aã]o\s*[:\-]\s*(.+)",
            r"Melhorar\s*(.+)",
            r"Verificar\s*(.+)",
            r"Ajustar\s*(.+)"
        ]
        
        for padrao in padroes:
            matches = re.findall(padrao, analise, re.IGNORECASE)
            recomendacoes.extend(matches)
        
        return recomendacoes[:5]  # Limitar a 5 recomendações

# =========================================================
#  PROCESSADOR EXCEL  - LÊ TODAS AS LINHAS)
# =========================================================

class ProcessadorExcel:
    def __init__(self):
        self.layouts_suportados = ['NF.xlsx', 'notas_fiscais']
        
    def extrair_dados_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extrai dados de notas fiscais de arquivo Excel
        """
        try:
            print(f"📊 Processando arquivo Excel: {os.path.basename(file_path)}")
            
            # Verificar se o arquivo existe
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
            
            # Carregar workbook
            workbook = load_workbook(file_path, data_only=True)
            print(f"✅ Workbook carregado. Planilhas: {workbook.sheetnames}")
            
            # Verificar se tem as planilhas esperadas
            if 'notas_fiscais' not in workbook.sheetnames:
                raise ValueError("Planilha 'notas_fiscais' não encontrada no arquivo Excel")
            
            # PRÉ-CARREGAR TODOS OS DADOS DAS PLANILHAS
            df_notas = pd.read_excel(file_path, sheet_name='notas_fiscais')
            print(f"📋 Encontradas {len(df_notas)} notas fiscais na planilha principal")
            
            # Carregar itens se existir
            df_itens = None
            if 'itens_nf' in workbook.sheetnames:
                df_itens = pd.read_excel(file_path, sheet_name='itens_nf')
                print(f"📦 Encontrados {len(df_itens)} itens na planilha de itens")
            
            # Carregar totais se existir
            df_totais = None
            if 'totais_nf' in workbook.sheetnames:
                df_totais = pd.read_excel(file_path, sheet_name='totais_nf')
                print(f"💰 Encontrados {len(df_totais)} registros na planilha de totais")
            
            workbook.close()
            
            resultados = []
            
            # PROCESSAR TODAS AS LINHAS DA PLANILHA DE NOTAS
            for idx, nota in df_notas.iterrows():
                try:
                    # Pular linhas vazias
                    if pd.isna(nota.get('id_nf')) or pd.isna(nota.get('numero')):
                        continue
                    
                    dados_nota = self._processar_linha_nota(nota, df_itens, df_totais, file_path)
                    if dados_nota:
                        resultados.append(dados_nota)
                        print(f"  ✅ Nota {nota.get('numero', 'N/A')} processada")
                    else:
                        print(f"  ⚠️  Nota {idx} não processada - dados insuficientes")
                        
                except Exception as e:
                    print(f"  ❌ Erro na nota {idx}: {e}")
                    continue
            
            print(f"✅ Processamento Excel concluído: {len(resultados)} notas extraídas")
            return resultados
            
        except Exception as e:
            logger.error(f"❌ Erro no processamento do Excel {file_path}: {e}")
            return []

    def _processar_linha_nota(self, nota: pd.Series, df_itens: pd.DataFrame, df_totais: pd.DataFrame, file_path: str) -> Optional[Dict[str, Any]]:
        """
        Processa uma linha da planilha de notas fiscais
        """
        try:
            # Extrair dados básicos da nota
            id_nf = nota.get('id_nf')
            if pd.isna(id_nf):
                return None
            
            # Buscar itens da nota
            itens = self._buscar_itens_nota(id_nf, df_itens)
            
            # Buscar totais da nota
            totais = self._buscar_totais_nota(id_nf, df_totais)
            
            # Construir estrutura de dados padronizada
            dados_padronizados = {
                # Identificação
                'numero': str(nota.get('numero', '')),
                'serie': str(nota.get('serie', '')),
                'data_emissao': self._converter_data_excel(nota.get('data_emissao')),
                'data_saida': self._converter_data_excel(nota.get('data_saida')),
                'natureza_operacao': nota.get('natureza_operacao', ''),
                'tipo_operacao': nota.get('tipo_operacao', ''),
                'finalidade_nf': nota.get('finalidade_nf', ''),
                
                # Emitente
                'cnpj_emitente': self._limpar_cnpj(nota.get('cnpj_emitente', '')),
                'razao_emitente': nota.get('razao_emitente', ''),
                
                # Destinatário
                'cnpj_destinatario': self._limpar_cnpj(nota.get('cnpj_destinatario', '')),
                'razao_destinatario': nota.get('razao_destinatario', ''),
                
                # Localidade
                'codigo_municipio': str(nota.get('codigo_municipio', '')),
                'uf_emitente': nota.get('uf_emitente', ''),
                'uf_destinatario': nota.get('uf_destinatario', ''),
                
                # Valores
                'valor_total': float(nota.get('valor_total', 0)),
                'valor_produtos': float(nota.get('valor_produtos', 0)),
                'valor_icms': float(nota.get('valor_icms', 0)),
                'valor_pis': float(nota.get('valor_pis', 0)),
                'valor_cofins': float(nota.get('valor_cofins', 0)),
                
                # Itens e totais
                'itens': itens,
                'totais_excel': totais,
                
                # Metadados
                'id_nf_excel': int(id_nf),
                'layout_identificado': 'EXCEL_NF',
                'metadados': {
                    'tipo_documento': 'NF_Excel',
                    'processado_em': datetime.now().isoformat(),
                    'arquivo_origem': os.path.basename(file_path)
                }
            }
            
            return dados_padronizados
            
        except Exception as e:
            logger.error(f"Erro ao processar linha da nota: {e}")
            return None

    def _buscar_itens_nota(self, id_nf: int, df_itens: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Busca itens da nota fiscal na planilha de itens
        """
        itens = []
        
        try:
            if df_itens is not None and not df_itens.empty:
                # Converter id_nf para o mesmo tipo que está no DataFrame
                id_nf_search = id_nf
                
                # Tentar diferentes formas de busca
                try:
                    itens_nota = df_itens[df_itens['id_nf'] == id_nf_search]
                except:
                    # Se falhar, tentar converter tipos
                    try:
                        itens_nota = df_itens[df_itens['id_nf'].astype(str) == str(id_nf_search)]
                    except:
                        itens_nota = pd.DataFrame()
                
                if not itens_nota.empty:
                    for _, item in itens_nota.iterrows():
                        # Pular itens vazios
                        if pd.isna(item.get('n_item')):
                            continue
                            
                        item_dict = {
                            'numero_item': int(item.get('n_item', 0)),
                            'codigo_produto': str(item.get('codigo_produto', '')),
                            'descricao': item.get('descricao_produto', ''),
                            'cfop': str(item.get('cfop', '')),
                            'unidade_comercial': item.get('unidade', ''),
                            'quantidade_comercial': float(item.get('quantidade', 0)),
                            'valor_unitario': float(item.get('valor_unitario', 0)),
                            'valor_total': float(item.get('valor_total', 0)),
                            'aliquota_icms': float(item.get('aliquota_icms', 0)),
                            'valor_icms_item': float(item.get('valor_icms', 0)),
                            'aliquota_pis': float(item.get('aliquota_pis', 0)),
                            'valor_pis_item': float(item.get('valor_pis', 0)),
                            'aliquota_cofins': float(item.get('aliquota_cofins', 0)),
                            'valor_cofins_item': float(item.get('valor_cofins', 0))
                        }
                        itens.append(item_dict)
                    
                    print(f"    📦 Encontrados {len(itens)} itens para NF {id_nf}")
                    
        except Exception as e:
            logger.warning(f"Erro ao buscar itens da nota {id_nf}: {e}")
            
        return itens

    def _buscar_totais_nota(self, id_nf: int, df_totais: pd.DataFrame) -> Dict[str, Any]:
        """
        Busca totais da nota fiscal na planilha de totais
        """
        totais = {}
        
        try:
            if df_totais is not None and not df_totais.empty:
                # Converter id_nf para o mesmo tipo que está no DataFrame
                id_nf_search = id_nf
                
                # Tentar diferentes formas de busca
                try:
                    total_nota = df_totais[df_totais['id_nf'] == id_nf_search]
                except:
                    try:
                        total_nota = df_totais[df_totais['id_nf'].astype(str) == str(id_nf_search)]
                    except:
                        total_nota = pd.DataFrame()
                
                if not total_nota.empty:
                    total = total_nota.iloc[0]
                    totais = {
                        'base_icms': float(total.get('base_icms', 0)),
                        'valor_icms': float(total.get('valor_icms', 0)),
                        'base_icms_st': float(total.get('base_icms_st', 0)),
                        'valor_icms_st': float(total.get('valor_icms_st', 0)),
                        'valor_produtos': float(total.get('valor_produtos', 0)),
                        'valor_frete': float(total.get('valor_frete', 0)),
                        'valor_seguro': float(total.get('valor_seguro', 0)),
                        'valor_desconto': float(total.get('valor_desconto', 0)),
                        'valor_outros': float(total.get('valor_outros', 0)),
                        'valor_pis': float(total.get('valor_pis', 0)),
                        'valor_cofins': float(total.get('valor_cofins', 0)),
                        'valor_nf': float(total.get('valor_nf', 0))
                    }
                    print(f"    💰 Totais encontrados para NF {id_nf}")
                    
        except Exception as e:
            logger.warning(f"Erro ao buscar totais da nota {id_nf}: {e}")
            
        return totais

    def _limpar_cnpj(self, cnpj: str) -> str:
        """Remove caracteres não numéricos do CNPJ"""
        if pd.isna(cnpj):
            return ''
        cnpj_limpo = re.sub(r'\D', '', str(cnpj))
        return cnpj_limpo if len(cnpj_limpo) == 14 else ''

    def _converter_data_excel(self, data_excel) -> Optional[str]:
        """Converte data do Excel para string ISO"""
        if pd.isna(data_excel):
            return None
        
        try:
            if isinstance(data_excel, (datetime, pd.Timestamp)):
                return data_excel.strftime('%Y-%m-%d')
            elif isinstance(data_excel, str):
                # Tentar parsear string de data
                try:
                    data_obj = datetime.strptime(data_excel, '%Y-%m-%d')
                    return data_obj.strftime('%Y-%m-%d')
                except:
                    return data_excel
            else:
                return str(data_excel)
        except:
            return None

# =========================================================
#  GESTOR DE BANCO DE DADOS
# =========================================================

class GestorBancoDados:
    def __init__(self):
        self.engine = None
        self.Session = None
        self._inicializar_banco()
    
    def _inicializar_banco(self):
        logger.info("🔄 Inicializando base de dados...")
        self.engine = self._tentar_mysql()
        
        if not self.engine:
            raise Exception("Não foi possível inicializar o banco de dados MySQL")
        
        self.Session = sessionmaker(bind=self.engine)
        self._criar_tabelas_verificar()
        logger.info("✅ Base de dados inicializada com sucesso!")
    
    def _tentar_mysql(self):
        try:
            if not all([MYSQL_CONFIG.get('user'), MYSQL_CONFIG.get('host')]):
                return None
            
            # Primeiro conectar sem especificar o banco para criá-lo se necessário
            engine_temp = create_engine(
                f"mysql+pymysql://{MYSQL_CONFIG['user']}:{MYSQL_CONFIG.get('password', '')}@{MYSQL_CONFIG['host']}:{MYSQL_CONFIG.get('port', 3306)}/",
                connect_args={'connect_timeout': 10}
            )
            
            with engine_temp.connect() as conn:
                # Criar banco se não existir
                conn.execute(text(f"CREATE DATABASE IF NOT EXISTS {MYSQL_CONFIG['database']}"))
                logger.info(f"✅ Banco de dados '{MYSQL_CONFIG['database']}' verificado/criado")
            
            # Agora conectar ao banco específico
            engine = create_engine(get_mysql_uri(), pool_pre_ping=True, echo=False)
            
            # Testar conexão
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            
            logger.info("✅ Conectado ao MySQL com sucesso!")
            return engine
            
        except Exception as e:
            logger.warning(f"❌ Não foi possível conectar ao MySQL: {e}")
            return None
    
    def _criar_tabelas_verificar(self):
        """Cria todas as tabelas e verifica se foram criadas"""
        try:
            logger.info("🔄 Criando/verificando tabelas...")
            
            # Criar todas as tabelas
            Base.metadata.create_all(self.engine)
            
            # Verificar tabelas criadas
            with self.engine.connect() as conn:
                result = conn.execute(text("SHOW TABLES"))
                tabelas = [row[0] for row in result]
                logger.info(f"✅ Tabelas no banco: {', '.join(tabelas)}")
            
            # Verificar tabelas essenciais
            tabelas_esperadas = ['emitentes', 'destinatarios', 'notas_fiscais', 'itens_nf']
            tabelas_faltantes = [t for t in tabelas_esperadas if t not in tabelas]
            
            if tabelas_faltantes:
                logger.warning(f"⚠️  Tabelas faltantes: {tabelas_faltantes}")
                # Tentar criar novamente
                for tabela in tabelas_faltantes:
                    try:
                        if tabela == 'emitentes':
                            Emitente.__table__.create(self.engine)
                        elif tabela == 'destinatarios':
                            Destinatario.__table__.create(self.engine)
                        elif tabela == 'notas_fiscais':
                            NotaFiscal.__table__.create(self.engine)
                        elif tabela == 'itens_nf':
                            ItemNF.__table__.create(self.engine)
                        logger.info(f"✅ Tabela {tabela} criada individualmente")
                    except Exception as e:
                        logger.error(f"❌ Erro ao criar tabela {tabela}: {e}")
            else:
                logger.info("✅ Todas as tabelas foram criadas com sucesso!")
                
        except Exception as e:
            logger.error(f"❌ Erro crítico ao criar tabelas: {e}")
            raise
    
    def testar_conexao(self):
        try:
            session = self.Session()
            resultado = session.execute(text("SELECT 1")).fetchone()
            session.close()
            return resultado[0] == 1
        except Exception as e:
            logger.error(f"Erro ao testar conexão: {e}")
            return False

# =========================================================
#  AGENTE DE EXTRAÇÃO FISCAL INTELIGENTE
# =========================================================

class AgenteExtracaoFiscalInteligente:
    def __init__(self):
        try:
            self.gestor_bd = GestorBancoDados()
            self.agente_inteligente = AgenteInteligenteDeepSeek()
            self.processador_excel = ProcessadorExcel()
            logger.info("✅ Agente de Extração Fiscal inicializado com sucesso!")
        except Exception as e:
            logger.error(f"❌ Erro ao inicializar Agente de Extração Fiscal: {e}")
            raise
    
    # MÉTODOS EXISTENTES PARA PROCESSAMENTO DE XML E ZIP (mantidos intactos)
    def _reparar_xml(self, content: str) -> str:
        """Repara XML corrompido ou mal formatado"""
        if content.startswith('\ufeff'):
            content = content[1:]
        
        content = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', content)
        content = content.replace('&', '&amp;')
        content = re.sub(r'&amp;(amp|lt|gt|quot|apos);', r'&\1;', content)
        
        if not content.strip().startswith('<?xml'):
            match = re.search(r'<[?]xml[^>]*>', content)
            if not match:
                content = '<?xml version="1.0" encoding="UTF-8"?>' + content
        
        return content

    def _reparar_xml_avancado(self, content: str) -> str:
        """Repara XML corrompido de forma mais agressiva"""
        content = content.replace('\ufeff', '').replace('\uFEFF', '')
        content = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', content)
        
        content = content.replace('&', '&amp;')
        content = re.sub(r'&amp;(amp|lt|gt|quot|apos);', r'&\1;', content)
        
        declarations = re.findall(r'<\?xml[^>]*\?>', content)
        if len(declarations) > 1:
            content = re.sub(r'<\?xml[^>]*\?>', '', content, count=len(declarations)-1)
        
        if not re.search(r'<\?xml[^>]*\?>', content):
            content = '<?xml version="1.0" encoding="UTF-8"?>' + content
        
        xml_declaration = re.search(r'<\?xml[^>]*\?>', content)
        if xml_declaration:
            content = content[xml_declaration.start():]
        
        root_match = re.search(r'<([^>?!\s][^>]*)>', content)
        if root_match:
            root_tag = root_match.group(1).split()[0]
            if f'</{root_tag}>' not in content:
                content = content + f'</{root_tag}>'
        
        return content
    
    def _ler_xml_com_tolerancia(self, xml_path: str) -> str:
        """Lê XML com tolerância a erros de encoding"""
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
        
        for encoding in encodings:
            try:
                with open(xml_path, 'r', encoding=encoding) as f:
                    content = f.read()
                return self._reparar_xml(content)
            except UnicodeDecodeError:
                continue
        
        try:
            with open(xml_path, 'rb') as f:
                content = f.read().decode('utf-8', errors='ignore')
            return self._reparar_xml(content)
        except Exception as e:
            raise Exception(f"Não foi possível ler o arquivo: {e}")

    def _estrategia_parsing_direto(self, xml_path: str) -> Dict[str, Any]:
        """Estratégia 1: Parsing direto padrão"""
        tree = ET.parse(xml_path)
        root = tree.getroot()
        dados = self._extrair_dados_basicos_xml(root)
        return self._salvar_nota_fiscal_simples(dados) if dados else {"erro": "Dados não extraídos"}

    def _estrategia_encoding_corrigido(self, xml_path: str) -> Dict[str, Any]:
        """Estratégia 2: Corrigir encoding e reparar XML"""
        content = self._ler_xml_com_tolerancia(xml_path)
        content = self._reparar_xml_avancado(content)
        root = ET.fromstring(content)
        dados = self._extrair_dados_basicos_xml(root)
        return self._salvar_nota_fiscal_simples(dados) if dados else {"erro": "Dados não extraídos"}

    def _estrategia_limpeza_agressiva(self, xml_path: str) -> Dict[str, Any]:
        """Estratégia 3: Limpeza agressiva do conteúdo XML"""
        with open(xml_path, 'rb') as f:
            raw_content = f.read()
        
        content_str = raw_content.decode('utf-8', errors='ignore')
        
        nfe_start = content_str.find('<NFe')
        if nfe_start == -1:
            nfe_start = content_str.find('<nfe')
        if nfe_start == -1:
            nfe_start = content_str.find('<NFE')
        
        if nfe_start != -1:
            nfe_end = content_str.find('</NFe>')
            if nfe_end == -1:
                nfe_end = content_str.find('</nfe>')
            if nfe_end == -1:
                nfe_end = content_str.find('</NFE>')
            
            if nfe_end != -1:
                nfe_end += 6
                content_str = content_str[nfe_start:nfe_end]
        
        content_str = self._reparar_xml_avancado(content_str)
        root = ET.fromstring(content_str)
        dados = self._extrair_dados_basicos_xml(root)
        return self._salvar_nota_fiscal_simples(dados) if dados else {"erro": "Dados não extraídos"}

    def _estrategia_extracao_minima(self, xml_path: str) -> Dict[str, Any]:
        """Estratégia 4: Extração mínima via regex como fallback final"""
        try:
            with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            numero_match = re.search(r'<nNF[^>]*>([^<]+)</nNF>', content, re.IGNORECASE)
            serie_match = re.search(r'<serie[^>]*>([^<]+)</serie>', content, re.IGNORECASE)
            razao_match = re.search(r'<xNome[^>]*>([^<]+)</xNome>', content, re.IGNORECASE)
            cnpj_match = re.search(r'<CNPJ[^>]*>([^<]+)</CNPJ>', content, re.IGNORECASE)
            valor_match = re.search(r'<vNF[^>]*>([^<]+)</vNF>', content, re.IGNORECASE)
            
            dados = {
                'numero': numero_match.group(1) if numero_match else '000000',
                'serie': serie_match.group(1) if serie_match else '1',
                'razao_emitente': razao_match.group(1) if razao_match else 'Emitente Não Identificado',
                'cnpj_emitente': cnpj_match.group(1) if cnpj_match else '00000000000000',
                'valor_total': float(valor_match.group(1)) if valor_match else 0.0,
                'data_emissao': datetime.now(),
                'modelo': '55'
            }
            
            return self._salvar_nota_fiscal_simples(dados)
            
        except Exception as e:
            return {"erro": f"Falha na extração mínima: {str(e)}"}
    
    def processar_xml_nfe(self, xml_path: str) -> Dict[str, Any]:
        """Processa XML com múltiplas estratégias de fallback"""
        try:
            logger.info(f"📄 Processando XML: {xml_path}")
            
            if not os.path.exists(xml_path):
                return {"erro": f"Arquivo não encontrado: {xml_path}"}
            
            strategies = [
                self._estrategia_parsing_direto,
                self._estrategia_encoding_corrigido, 
                self._estrategia_limpeza_agressiva,
                self._estrategia_extracao_minima
            ]
            
            for i, strategy in enumerate(strategies):
                try:
                    logger.info(f"🔄 Tentando estratégia {i+1} para {xml_path}")
                    resultado = strategy(xml_path)
                    if resultado and not resultado.get('erro'):
                        logger.info(f"✅ Estratégia {i+1} funcionou para {xml_path}")
                        logger.info(f"📊 Dados extraídos: {resultado.get('numero', 'N/A')} - {resultado.get('razao_emitente', 'N/A')}")
                        return resultado
                    else:
                        logger.warning(f"⚠️ Estratégia {i+1} retornou erro: {resultado.get('erro', 'Erro desconhecido')}")
                except Exception as e:
                    logger.warning(f"❌ Estratégia {i+1} falhou: {e}")
                    continue
            
            return {"erro": "Todas as estratégias de parsing falharam"}
            
        except Exception as e:
            logger.error(f"❌ Erro crítico ao processar XML: {e}")
            return {"erro": f"Falha crítica no processamento: {str(e)}"}
    
    def _extrair_dados_basicos_xml(self, root) -> Dict[str, Any]:
        """Extrai dados básicos do XML com tolerância a erros"""
        try:
            namespaces = [
                'http://www.portalfiscal.inf.br/nfe',
                'http://www.portalfiscal.inf.br/nfe/3.10',
                'http://www.portalfiscal.inf.br/nfe/4.00',
                None
            ]
            
            for ns in namespaces:
                try:
                    if ns:
                        ns_prefix = f'{{{ns}}}'
                    else:
                        ns_prefix = ''
                    
                    ide = root.find(f'.//{ns_prefix}ide')
                    emit = root.find(f'.//{ns_prefix}emit')
                    total = root.find(f'.//{ns_prefix}total')
                    
                    if ide is not None:
                        break
                except:
                    continue
            
            if ide is None:
                for elem in root.iter():
                    tag = elem.tag.lower()
                    if 'ide' in tag:
                        ide = elem
                    if 'emit' in tag:
                        emit = elem
                    if 'total' in tag:
                        total = elem
            
            numero = None
            serie = None
            razao_social = None
            cnpj = None
            valor_total = 0.0
            
            if ide is not None:
                numero_elem = ide.find('nNF') or ide.find('nNF'.lower())
                serie_elem = ide.find('serie') or ide.find('serie'.lower())
                
                if numero_elem is not None:
                    numero = numero_elem.text
                if serie_elem is not None:
                    serie = serie_elem.text
            
            if emit is not None:
                razao_elem = emit.find('xNome') or emit.find('xNome'.lower()) or emit.find('xnome')
                cnpj_elem = emit.find('CNPJ') or emit.find('CNPJ'.lower()) or emit.find('cnpj')
                
                if razao_elem is not None:
                    razao_social = razao_elem.text
                if cnpj_elem is not None:
                    cnpj = cnpj_elem.text
            
            if total is not None:
                icms_total = total.find('ICMSTot') or total.find('ICMSTot'.lower()) or total.find('icmstot')
                if icms_total is not None:
                    valor_elem = icms_total.find('vNF') or icms_total.find('vNF'.lower()) or icms_total.find('vnf')
                    if valor_elem is not None:
                        try:
                            valor_total = float(valor_elem.text or 0)
                        except:
                            valor_total = 0.0
            
            return {
                'numero': numero or '000000',
                'serie': serie or '1',
                'razao_emitente': razao_social or 'Emitente Não Identificado',
                'cnpj_emitente': cnpj or '00000000000000',
                'valor_total': valor_total,
                'data_emissao': datetime.now(),
                'modelo': '55'
            }
            
        except Exception as e:
            logger.error(f"Erro ao extrair dados básicos: {e}")
            return {
                'numero': '000000',
                'serie': '1',
                'razao_emitente': 'Emitente Não Identificado',
                'cnpj_emitente': '00000000000000',
                'valor_total': 0.0,
                'data_emissao': datetime.now(),
                'modelo': '55'
            }
    
    def _salvar_nota_fiscal_simples(self, dados_nota: Dict[str, Any]) -> Dict[str, Any]:
        """Salva nota fiscal de forma simplificada com gestão adequada de sessão"""
        session = self.gestor_bd.Session()
        try:
            # Garantir que o emitente existe
            emitente = session.query(Emitente).filter_by(cnpj=dados_nota['cnpj_emitente']).first()
            if not emitente:
                emitente = Emitente(
                    cnpj=dados_nota['cnpj_emitente'],
                    razao_social=dados_nota['razao_emitente']
                )
                session.add(emitente)
                session.flush()

            # Verificar se a nota já existe
            nota_existente = session.query(NotaFiscal).filter_by(
                numero=dados_nota['numero'],
                serie=dados_nota.get('serie', ''),
                cnpj_emitente=dados_nota.get('cnpj_emitente', '')
            ).first()
            
            if nota_existente:
                for key, value in dados_nota.items():
                    if hasattr(nota_existente, key) and value is not None:
                        setattr(nota_existente, key, value)
                nota_existente.updated_at = datetime.now()
                resultado = {**dados_nota, 'id_nf': nota_existente.id_nf, 'atualizado': True}
            else:
                nova_nota = NotaFiscal(**dados_nota)
                session.add(nova_nota)
                session.flush()
                resultado = {**dados_nota, 'id_nf': nova_nota.id_nf, 'novo': True}
            
            session.commit()
            logger.info(f"✅ Nota {dados_nota.get('numero')} salva com sucesso")
            return resultado
            
        except Exception as e:
            session.rollback()
            logger.error(f"❌ Erro ao salvar nota fiscal: {e}")
            return {"erro": str(e)}
        finally:
            session.close()
    
    def processar_zip_nf(self, zip_path: str) -> List[Dict[str, Any]]:
        """Processa arquivo ZIP com tratamento robusto de erros"""
        try:
            logger.info(f"Processando ZIP: {zip_path}")
            resultados = []
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                all_files = zip_ref.namelist()
                xml_files = [f for f in all_files if f.lower().endswith('.xml')]
                
                if not xml_files:
                    return [{"erro": "Nenhum arquivo XML encontrado no ZIP", "tipo": "zip_vazio"}]
                
                logger.info(f"Encontrados {len(xml_files)} arquivos XML no ZIP")
                
                for xml_file in xml_files:
                    try:
                        logger.info(f"📄 Processando: {xml_file}")
                        
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as tmp_file:
                            file_content = zip_ref.read(xml_file)
                            tmp_file.write(file_content)
                            tmp_path = tmp_file.name
                        
                        resultado = self.processar_xml_nfe(tmp_path)
                        resultado['arquivo'] = xml_file
                        resultados.append(resultado)
                        
                        try:
                            os.unlink(tmp_path)
                        except:
                            pass
                            
                    except Exception as e:
                        logger.error(f"❌ Erro ao processar {xml_file}: {e}")
                        resultados.append({
                            "erro": str(e), 
                            "arquivo": xml_file,
                            "tipo": "processamento_arquivo"
                        })
            
            sucessos = [r for r in resultados if 'id_nf' in r or 'numero' in r]
            erros = [r for r in resultados if 'erro' in r]
            
            logger.info(f"✅ ZIP processado: {len(sucessos)} sucessos, {len(erros)} erros")
            return resultados
            
        except zipfile.BadZipFile:
            erro_msg = "Arquivo ZIP corrompido ou inválido"
            logger.error(f"❌ {erro_msg}")
            return [{"erro": erro_msg, "tipo": "zip_corrompido"}]
        except Exception as e:
            logger.error(f"❌ Erro ao processar ZIP: {e}")
            return [{"erro": str(e), "tipo": "processamento_zip"}]
    
    def processar_excel_nf(self, excel_path: str) -> List[Dict[str, Any]]:
        """Processa arquivo Excel usando o ProcessadorExcel"""
        try:
            logger.info(f"📊 Processando Excel: {excel_path}")
            
            # Usar o processador de Excel para extrair os dados
            dados_notas = self.processador_excel.extrair_dados_excel(excel_path)
            
            resultados = []
            for dados_nota in dados_notas:
                try:
                    # Salvar a nota completa (com itens e totais)
                    resultado = self._salvar_nota_fiscal_completa(dados_nota)
                    resultados.append(resultado)
                    
                    # Opcional: Analisar com agente inteligente
                    if hasattr(self, 'agente_inteligente'):
                        analise = self.agente_inteligente.analisar_dados_fiscais(dados_nota, "nota_fiscal")
                        if analise.get('analisado'):
                            resultado['analise_inteligente'] = analise
                            
                except Exception as e:
                    logger.error(f"Erro ao salvar nota do Excel: {e}")
                    resultados.append({"erro": str(e), "dados": dados_nota.get('numero', 'N/A')})
            
            logger.info(f"✅ Excel processado: {len(resultados)} registros")
            return resultados
            
        except Exception as e:
            logger.error(f"❌ Erro ao processar Excel: {e}")
            return [{"erro": str(e)}]
    
    def _salvar_nota_fiscal_completa(self, dados_nota: Dict[str, Any]) -> Dict[str, Any]:
        """Salva nota fiscal completa com itens e totais"""
        session = self.gestor_bd.Session()
        try:
            # 1. Verificar/Criar Emitente
            emitente = session.query(Emitente).filter_by(cnpj=dados_nota['cnpj_emitente']).first()
            if not emitente:
                emitente = Emitente(
                    cnpj=dados_nota['cnpj_emitente'],
                    razao_social=dados_nota['razao_emitente']
                )
                session.add(emitente)
                session.flush()

            # 2. Verificar/Criar Destinatário (se houver)
            destinatario = None
            if dados_nota.get('cnpj_destinatario'):
                destinatario = session.query(Destinatario).filter_by(cnpj=dados_nota['cnpj_destinatario']).first()
                if not destinatario:
                    destinatario = Destinatario(
                        cnpj=dados_nota['cnpj_destinatario'],
                        razao_social=dados_nota.get('razao_destinatario', '')
                    )
                    session.add(destinatario)
                    session.flush()

            # 3. Verificar se a nota já existe
            nota_existente = session.query(NotaFiscal).filter_by(
                numero=dados_nota['numero'],
                serie=dados_nota.get('serie', ''),
                cnpj_emitente=dados_nota['cnpj_emitente']
            ).first()
            
            if nota_existente:
                # Atualizar nota existente
                for key, value in dados_nota.items():
                    if hasattr(nota_existente, key) and value is not None and key not in ['itens', 'totais_excel']:
                        setattr(nota_existente, key, value)
                nota_existente.updated_at = datetime.now()
                nota = nota_existente
                resultado = {**dados_nota, 'id_nf': nota.id_nf, 'atualizado': True}
            else:
                # Criar nova nota
                nota = NotaFiscal(
                    numero=dados_nota['numero'],
                    serie=dados_nota.get('serie', ''),
                    data_emissao=dados_nota.get('data_emissao'),
                    data_saida=dados_nota.get('data_saida'),
                    natureza_operacao=dados_nota.get('natureza_operacao'),
                    tipo_operacao=dados_nota.get('tipo_operacao'),
                    finalidade_nf=dados_nota.get('finalidade_nf'),
                    valor_total=dados_nota.get('valor_total'),
                    valor_produtos=dados_nota.get('valor_produtos'),
                    valor_icms=dados_nota.get('valor_icms'),
                    valor_pis=dados_nota.get('valor_pis'),
                    valor_cofins=dados_nota.get('valor_cofins'),
                    cnpj_emitente=dados_nota['cnpj_emitente'],
                    razao_emitente=dados_nota['razao_emitente'],
                    cnpj_destinatario=dados_nota.get('cnpj_destinatario'),
                    razao_destinatario=dados_nota.get('razao_destinatario'),
                    codigo_municipio=dados_nota.get('codigo_municipio'),
                    uf_emitente=dados_nota.get('uf_emitente'),
                    uf_destinatario=dados_nota.get('uf_destinatario'),
                    layout_detectado=dados_nota.get('layout_identificado'),
                    tipo_arquivo='EXCEL'
                )
                session.add(nota)
                session.flush()
                resultado = {**dados_nota, 'id_nf': nota.id_nf, 'novo': True}

            # 4. Salvar itens
            if 'itens' in dados_nota and dados_nota['itens']:
                # Remover itens existentes (se estiver atualizando)
                session.query(ItemNF).filter_by(id_nf=nota.id_nf).delete()
                
                for item_data in dados_nota['itens']:
                    item = ItemNF(
                        id_nf=nota.id_nf,
                        n_item=item_data['numero_item'],
                        codigo_produto=item_data.get('codigo_produto'),
                        descricao_produto=item_data.get('descricao'),
                        cfop=item_data.get('cfop'),
                        quantidade=item_data.get('quantidade_comercial'),
                        unidade=item_data.get('unidade_comercial'),
                        valor_unitario=item_data.get('valor_unitario'),
                        valor_total=item_data.get('valor_total'),
                        aliquota_icms=item_data.get('aliquota_icms'),
                        valor_icms=item_data.get('valor_icms_item'),
                        aliquota_pis=item_data.get('aliquota_pis'),
                        valor_pis=item_data.get('valor_pis_item'),
                        aliquota_cofins=item_data.get('aliquota_cofins'),
                        valor_cofins=item_data.get('valor_cofins_item')
                    )
                    session.add(item)

            # 5. Salvar totais (se existirem)
            if 'totais_excel' in dados_nota and dados_nota['totais_excel']:
                # Remover totais existentes
                session.query(TotalNF).filter_by(id_nf=nota.id_nf).delete()
                
                totais_data = dados_nota['totais_excel']
                total = TotalNF(
                    id_nf=nota.id_nf,
                    base_icms=totais_data.get('base_icms'),
                    valor_icms=totais_data.get('valor_icms'),
                    base_icms_st=totais_data.get('base_icms_st'),
                    valor_icms_st=totais_data.get('valor_icms_st'),
                    valor_produtos=totais_data.get('valor_produtos'),
                    valor_frete=totais_data.get('valor_frete'),
                    valor_seguro=totais_data.get('valor_seguro'),
                    valor_desconto=totais_data.get('valor_desconto'),
                    valor_outros=totais_data.get('valor_outros'),
                    valor_pis=totais_data.get('valor_pis'),
                    valor_cofins=totais_data.get('valor_cofins'),
                    valor_nf=totais_data.get('valor_nf')
                )
                session.add(total)

            session.commit()
            logger.info(f"✅ Nota {dados_nota.get('numero')} salva com sucesso (completa)")
            return resultado
            
        except Exception as e:
            session.rollback()
            logger.error(f"❌ Erro ao salvar nota fiscal completa: {e}")
            return {"erro": str(e)}
        finally:
            session.close()
    
    def listar_documentos(self, limite: int = 100) -> List[Dict[str, Any]]:
        """Lista documentos processados com gestão adequada de sessão"""
        session = self.gestor_bd.Session()
        try:
            notas = session.query(NotaFiscal).order_by(
                NotaFiscal.data_emissao.desc()
            ).limit(limite).all()
            
            return [
                {
                    'id_nf': nota.id_nf,
                    'numero': nota.numero,
                    'serie': nota.serie,
                    'data_emissao': nota.data_emissao.isoformat() if nota.data_emissao else None,
                    'razao_emitente': nota.razao_emitente,
                    'valor_total': nota.valor_total,
                    'uf_emitente': nota.uf_emitente,
                    'cnpj_emitente': nota.cnpj_emitente
                }
                for nota in notas
            ]
        except Exception as e:
            logger.error(f"Erro ao listar documentos: {e}")
            return []
        finally:
            session.close()
    
    def analisar_lote_com_inteligencia(self, dados_lote: Dict[str, Any]) -> Dict[str, Any]:
        """Analisa um lote de processamento usando o agente inteligente"""
        if hasattr(self, 'agente_inteligente'):
            return self.agente_inteligente.analisar_dados_fiscais(dados_lote, "lote_processamento")
        else:
            return {"analisado": False, "erro": "Agente inteligente não disponível"}

# Teste
if __name__ == "__main__":
    try:
        agente = AgenteExtracaoFiscalInteligente()
        print("✅ Sistema fiscal inicializado!")
        
        docs = agente.listar_documentos()
        print(f"📄 Documentos no banco: {len(docs)}")
        
    except Exception as e:
        print(f"❌ Erro: {e}")